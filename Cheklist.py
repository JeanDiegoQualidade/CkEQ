import flet as ft
from openpyxl import load_workbook
from datetime import datetime

# Função para carregar dados da planilha
def carregar_dados():
    wb = load_workbook('BD.xlsx')
    ws_operadores = wb['Operadores']
    ws_chassis = wb['Chassis']
    ws_id = wb['ID']

    # Dados dos operadores
    dados = [ws_operadores.cell(row=i, column=1).value for i in range(2, ws_operadores.max_row + 1)]
    dados_completos = {}
    for i in range(2, ws_operadores.max_row + 1):
        nome = ws_operadores.cell(row=i, column=1).value
        senha = ws_operadores.cell(row=i, column=3).value
        valor_coluna_4 = ws_operadores.cell(row=i, column=4).value
        nivel = ws_operadores.cell(row=i, column=2).value  # Armazena o valor da coluna 2
        
        dados_completos[nome] = {'senha': senha, 'valor_coluna_4': valor_coluna_4, 'nivel': nivel}

    # Dados dos chassis
    fornecedores = [ws_chassis.cell(row=i, column=8).value for i in range(2, ws_chassis.max_row + 1)]
    ultimo_numero = ws_chassis.cell(row=1, column=7).value

    # Dados da planilha ID
    opcoes_id = [ws_id.cell(row=1, column=i).value for i in range(1, ws_id.max_column + 1)]  # Correção para ler até a última coluna
    dados_id = {opcao: [ws_id.cell(row=j, column=i).value for j in range(2, ws_id.max_row + 1)] for i, opcao in enumerate(opcoes_id, start=1)}

    return dados, dados_completos, fornecedores, ultimo_numero, opcoes_id, dados_id

# Função para verificar senha e gerar número de chassi
def verificar_senha(nome, senha, dados_completos, ultimo_numero):
    usuario = dados_completos.get(nome, {})
    senha_correta = str(usuario.get('senha', ''))
    
    # Remove espaços extras e compara as senhas
    if senha_correta.strip() != str(senha).strip():
        return False, None, None

    # Gerar o número do chassi
    semana_atual = datetime.now().strftime('%U')  # Número da semana atual
    ano_atual = datetime.now().strftime('%y')     # Dois últimos dígitos do ano atual
    numero_incrementado = int(ultimo_numero) + 1
    numero_chassi = f"{semana_atual}{ano_atual}{numero_incrementado:04d}"
        
    return True, numero_chassi, usuario.get('nivel', 1)  # Adiciona nível ao retorno

# Função para nível 2
def nivel2():
    print("Função nível 2 executada.")

# Função para nível maior que 2
def nivelmaior():
    print("Função nível maior que 2 executada.")

def main(page: ft.Page):
    page.title = "Aplicativo Flet"
    page.vertical_alignment = ft.MainAxisAlignment.START
    page.horizontal_alignment = ft.CrossAxisAlignment.START

    dados_operadores, dados_completos, fornecedores, ultimo_numero, opcoes_id, dados_id = carregar_dados()

    lista_suspensa = ft.Dropdown(
        options=[ft.dropdown.Option(text=dado, key=dado) for dado in dados_operadores if dado],
        label="Inspetor",
        width=300  # Definindo uma largura fixa para alinhamento consistente
    )

    campo_senha = ft.TextField(
        label="Senha",
        password=True,
        width=300  # Definindo uma largura fixa para alinhamento consistente
    )

    lista_id = ft.Dropdown(
        options=[ft.dropdown.Option(text=opcao, key=opcao) for opcao in opcoes_id if opcao],
        label="Equipamento",
        width=300
    )

    lista_detalhes = ft.Dropdown(
        options=[],  # Inicialmente vazio
        label="Modelo",
        width=300
    )

    dropdown_fornecedor = ft.Dropdown(
    options=[ft.dropdown.Option(text=fornecedor, key=fornecedor) for fornecedor in fornecedores if fornecedor],
    label="Fornecedor",
    width=300,  # Definindo uma largura fixa para alinhamento consistente
   )


    def atualizar_lista_detalhes(e):
        opcao_selecionada = lista_id.value
        if opcao_selecionada and opcao_selecionada in dados_id:
            lista_detalhes.options = [ft.dropdown.Option(text=item, key=item) for item in dados_id[opcao_selecionada] if item]
        else:
            lista_detalhes.options = []  # Limpa a lista se a opção não estiver em dados_id
        page.update()

    def ao_clicar_entrar(e):
        nome_selecionado = lista_suspensa.value
        senha_inserida = campo_senha.value

        senha_valida, numero_chassi, nivel = verificar_senha(nome_selecionado, senha_inserida, dados_completos, ultimo_numero)

        if senha_valida:
            # Verificar o nível e chamar a função correspondente
            if nivel == 1:
                page.controls.clear()
                page.add(
                    ft.Column(
                        controls=[
                            ft.ElevatedButton(text="Reiniciar", on_click=reiniciar),  # Botão Reiniciar na parte superior
                            ft.Text(value=f"Inspetor: {nome_selecionado}", size=16, text_align=ft.TextAlign.LEFT),
                            ft.Text(value=f"Tipo de Inspeção: {dados_completos[nome_selecionado]['valor_coluna_4']}", size=16, text_align=ft.TextAlign.LEFT),
                            ft.Text(value=f"Chassi: {numero_chassi}", size=16, text_align=ft.TextAlign.LEFT),  # Exibe o número do chassi
                            lista_id,
                            lista_detalhes,
                            ft.Dropdown(
                                options=[ft.dropdown.Option(text=fornecedor, key=fornecedor) for fornecedor in fornecedores if fornecedor],
                                label="Fornecedor",
                                width=300  # Definindo uma largura fixa para alinhamento consistente
                            ),
                            ft.ElevatedButton(text="Iniciar Inspeção", on_click=iniciar_inspecao)  # Botão Iniciar Inspeção
                        ],
                        horizontal_alignment=ft.CrossAxisAlignment.START,  # Alinha a coluna à esquerda
                        spacing=10  # Adiciona espaçamento entre os controles
                    )
                )

                lista_id.on_change = atualizar_lista_detalhes  # Atualiza a lista_detalhes quando a lista_id muda
                
            elif nivel == 2:
                nivel2()
            elif nivel > 2:
                nivelmaior()
        else:
            page.add(
                ft.Column(
                    controls=[
                        ft.ElevatedButton(text="Reiniciar", on_click=reiniciar),  # Botão Reiniciar na parte superior
                        ft.Text(value="Senha incorreta!", color="red", text_align=ft.TextAlign.LEFT)
                    ],
                    horizontal_alignment=ft.CrossAxisAlignment.START,  # Alinha a coluna à esquerda
                    spacing=10  # Adiciona espaçamento entre os controles
                )
            )

        page.update()

    def iniciar_inspecao(e):
        if not lista_suspensa.value or not lista_id.value or not lista_detalhes.value:
            page.add(
                ft.Column(
                    controls=[
                        ft.Text(value="Por favor, preencha todos os campos antes de iniciar a inspeção.", color="red", text_align=ft.TextAlign.LEFT),
                        ft.ElevatedButton(text="Reiniciar", on_click=reiniciar)  # Botão para reiniciar
                    ],
                    horizontal_alignment=ft.CrossAxisAlignment.START,  # Alinha a coluna à esquerda
                    spacing=10  # Adiciona espaçamento entre os controles
                )
            )
            page.update()
            return

        # Armazenar os valores em variáveis
        inspetor = lista_suspensa.value
        numerochassi = verificar_senha(inspetor, campo_senha.value, dados_completos, ultimo_numero)[1]
        fornecedor = dropdown_fornecedor.value  # O valor do fornecedor é o selecionado no dropdown

        wb = load_workbook('BD.xlsx')
        ws_inspecao = wb['InspecaoChassi']

        # Encontrar a coluna correspondente ao modelo
        coluna_modelo = None
        for col in ws_inspecao.iter_cols(min_row=1, max_row=1):
            if col[0].value == lista_detalhes.value:
                coluna_modelo = col[0].column
                break

        if not coluna_modelo:
            page.add(
                ft.Column(
                    controls=[
                        ft.Text(value="Modelo não encontrado na planilha de inspeção.", color="red", text_align=ft.TextAlign.LEFT),
                        ft.ElevatedButton(text="Reiniciar", on_click=reiniciar)  # Botão para reiniciar
                    ],
                    horizontal_alignment=ft.CrossAxisAlignment.START,  # Alinha a coluna à esquerda
                    spacing=10  # Adiciona espaçamento entre os controles
                )
            )
            page.update()
            return
########################################################################################################################################################################################
        # Criar o checklist
        
        def criar_checklist(pagina_atual):
            max_linhas_por_pagina = 5  # Ajuste o número de linhas por página conforme necessário
            controles = []

            for i, row in enumerate(ws_inspecao.iter_rows(min_row=2, max_row=ws_inspecao.max_row, min_col=1, max_col=ws_inspecao.max_column)):
                if row[coluna_modelo - 1].value == "x":
                    label = row[0].value
                    controles.append(
                        ft.Column(
                            controls=[
                                ft.Text(value=label, width=200),  # Ajustar a largura conforme necessário
                                ft.Row(
                                    controls=[
                                        ft.Checkbox(label="Ok"),
                                        ft.Checkbox(label="Não conforme")
                                    ],
                                    spacing=10
                                )
                            ],
                            alignment=ft.MainAxisAlignment.START,
                            spacing=5  # Adiciona espaçamento entre o texto e os checkboxes
                        )
                    )

            # Determinar o número de páginas necessárias
            num_paginas = -(-len(controles) // max_linhas_por_pagina)  # Ceil

            # Criar a página atual
            inicio = pagina_atual * max_linhas_por_pagina
            fim = inicio + max_linhas_por_pagina
            controles_pagina = controles[inicio:fim]

            page.controls.clear()
            page.add(
                ft.Column(
                    controls=[
                        ft.ElevatedButton(text="Reiniciar", on_click=reiniciar),  # Botão Reiniciar na parte superior
                        ft.ElevatedButton(text="Finalizar Inspeção", on_click=finalizar_inspecao),  # Botão Finalizar Inspeção
                        *controles_pagina,
                        ft.Row(
                            controls=[
                                ft.ElevatedButton(text="Página Anterior", on_click=lambda e: criar_checklist(pagina_atual - 1)) if pagina_atual > 0 else ft.Container(),
                                ft.Text(value=f"Página {pagina_atual + 1} de {num_paginas}"),
                                ft.ElevatedButton(text="Próxima Página", on_click=lambda e: criar_checklist(pagina_atual + 1)) if pagina_atual < num_paginas - 1 else ft.Container()
                            ],
                            alignment=ft.MainAxisAlignment.CENTER,
                            spacing=10
                        )
                    ],
                    horizontal_alignment=ft.CrossAxisAlignment.START,  # Alinha a coluna à esquerda
                    spacing=10  # Adiciona espaçamento entre os controles
                )
            )
            page.update()

        criar_checklist(pagina_atual=0)  # Iniciar com a primeira página
    
    
########################################################################################################################################################################
    def finalizar_inspecao(e):
        mensagens = []
        # Flag para verificar se pelo menos um checkbox foi selecionado
        pelo_menos_um_selecionado = False

        # Verificar se todos os checkboxes têm exatamente um selecionado
        for row in page.controls:
            if isinstance(row, ft.Row):
                checkboxes = [control for control in row.controls if isinstance(control, ft.Checkbox)]
                if len(checkboxes) == 2:
                    ok_selecionado = checkboxes[0].value
                    nao_conforme_selecionado = checkboxes[1].value
                    
                    # Verifica se pelo menos um checkbox está selecionado
                    if ok_selecionado or nao_conforme_selecionado:
                        pelo_menos_um_selecionado = True

                    if ok_selecionado and nao_conforme_selecionado:
                        mensagens.append("Você não pode selecionar 'Ok' e 'Não conforme' ao mesmo tempo.")
                    elif not ok_selecionado and not nao_conforme_selecionado:
                        mensagens.append("Por favor, selecione uma opção entre 'Ok' e 'Não conforme'.")
        
        # Verificar se pelo menos um checkbox foi selecionado
        if not pelo_menos_um_selecionado:
            mensagens.append("Pelo menos um checkbox deve ser selecionado.")

        if mensagens:
            page.add(
                ft.Column(
                    controls=[
                        ft.Text(value="\n".join(mensagens), color="red", text_align=ft.TextAlign.LEFT),
                        
                    ],
                    horizontal_alignment=ft.CrossAxisAlignment.START,  # Alinha a coluna à esquerda
                    spacing=10  # Adiciona espaçamento entre os controles
                )
            )
            page.update()
            return

        # Se tudo estiver OK
        page.add(
            ft.Column(
                controls=[
                    ft.Text(value="Inspeção concluída com sucesso!", color="green", text_align=ft.TextAlign.LEFT),
                    
                ],
                horizontal_alignment=ft.CrossAxisAlignment.START,  # Alinha a coluna à esquerda
                spacing=10  # Adiciona espaçamento entre os controles
            )
        )
        page.update()

        # Reiniciar o sistema (limpar controles e voltar à tela inicial)
        reiniciar(e)
        

####################################################################################################
    def reiniciar(e):
        # Limpar a tela e restaurar controles iniciais
        page.controls.clear()
        page.add(
            ft.Column(
                controls=[
                    ft.ElevatedButton(text="Reiniciar", on_click=reiniciar),  # Botão Reiniciar na parte superior
                    lista_suspensa,
                    campo_senha,
                    ft.ElevatedButton(text="Entrar", on_click=ao_clicar_entrar)
                ],
                horizontal_alignment=ft.CrossAxisAlignment.START,  # Alinha a coluna à esquerda
                spacing=10  # Adiciona espaçamento entre os controles
            )
        )
        campo_senha.value = ''
        page.update()

    page.add(
        ft.Column(
            controls=[
                ft.ElevatedButton(text="Reiniciar", on_click=reiniciar),  # Botão Reiniciar na parte superior
                lista_suspensa,
                campo_senha,
                ft.ElevatedButton(text="Entrar", on_click=ao_clicar_entrar)
            ],
            horizontal_alignment=ft.CrossAxisAlignment.START,  # Alinha a coluna à esquerda
            spacing=10  # Adiciona espaçamento entre os controles
        )
    )

# Executar o aplicativo
ft.app(target=main)
